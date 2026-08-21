from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File, Form, Query, Body
from fastapi.responses import Response, StreamingResponse
from fastapi.concurrency import run_in_threadpool
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from db_pg import Database
import os
import io
import asyncio
import logging
import uuid
import requests
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional
from datetime import datetime, timezone, timedelta
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

MADRID_TZ = ZoneInfo("Europe/Madrid")


def now_madrid() -> datetime:
    return datetime.now(MADRID_TZ)


def to_madrid(dt: Optional[datetime]) -> Optional[datetime]:
    if dt is None:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(MADRID_TZ)


# Postgres (Supabase) connection. `db` expone la misma forma que la `db` de Motor
# (db.naves.find(...), db.turnos.update_one(...), etc.) vía el adaptador en db_pg.py.
# El pool real de conexiones se crea de forma async en el evento de startup.
DATABASE_URL = os.environ['DATABASE_URL']
db = Database()

# ---------------- Emergent Object Storage ----------------
STORAGE_BASE = (os.environ.get("INTEGRATION_PROXY_URL") or "").strip() or "https://integrations.emergentagent.com"
STORAGE_URL = STORAGE_BASE.rstrip("/") + "/objstore/api/v1/storage"
EMERGENT_KEY = os.environ.get("EMERGENT_LLM_KEY")
APP_NAME = "control-guardia"
storage_key: Optional[str] = None


def init_storage():
    global storage_key
    if storage_key:
        return storage_key
    resp = requests.post(f"{STORAGE_URL}/init", json={"emergent_key": EMERGENT_KEY}, timeout=30)
    resp.raise_for_status()
    storage_key = resp.json()["storage_key"]
    return storage_key


def put_object(path: str, data: bytes, content_type: str) -> dict:
    global storage_key
    key = init_storage()
    resp = requests.put(
        f"{STORAGE_URL}/objects/{path}",
        headers={"X-Storage-Key": key, "Content-Type": content_type},
        data=data,
        timeout=120,
    )
    if resp.status_code == 503:
        storage_key = None
        key = init_storage()
        resp = requests.put(
            f"{STORAGE_URL}/objects/{path}",
            headers={"X-Storage-Key": key, "Content-Type": content_type},
            data=data,
            timeout=120,
        )
    resp.raise_for_status()
    return resp.json()


def get_object(path: str) -> tuple[bytes, str]:
    global storage_key
    key = init_storage()
    resp = requests.get(
        f"{STORAGE_URL}/objects/{path}",
        headers={"X-Storage-Key": key},
        timeout=60,
    )
    if resp.status_code == 503:
        storage_key = None
        key = init_storage()
        resp = requests.get(
            f"{STORAGE_URL}/objects/{path}",
            headers={"X-Storage-Key": key},
            timeout=60,
        )
    resp.raise_for_status()
    return resp.content, resp.headers.get("Content-Type", "application/octet-stream")


# ---------------- FastAPI ----------------
app = FastAPI()
api_router = APIRouter(prefix="/api")


# ---------------- Models ----------------
EVENT_TYPES = [
    "entrada", "salida", "ronda_inicio", "ronda_fin", "tarea", "incidencia", "descanso_inicio", "descanso_fin",
    "entrada_nave", "salida_nave", "llamada_centralita", "chequeo", "accion_nave",
    "vehiculo_vandalizado", "vehiculo_reparado",
]

SERVICES = ["Cándido Zamora"]
VEHICLE_TIPOS = ["Camión", "Grúa", "Contenedor", "Furgoneta", "Otro"]
TURNO_TIPOS = ["dia", "noche"]
AUTOFINALIZE_GRACE_MINUTES = 90  # margen de cortesía tras la hora de fin programada antes de auto-finalizar
LEGACY_TURNO_MAX_HOURS = 16  # turnos legacy (sin scheduled_end) se finalizan si llevan más de esto activos


class Nave(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    address: Optional[str] = None
    notes: Optional[str] = None
    service_name: Optional[str] = None
    has_access_buttons: bool = True
    check_items: List[str] = []
    custom_actions: List[str] = []
    has_vehicles: bool = False
    order: int = 0
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class NaveCreate(BaseModel):
    name: str
    address: Optional[str] = None
    notes: Optional[str] = None
    service_name: Optional[str] = None
    has_access_buttons: bool = True
    check_items: List[str] = []
    custom_actions: List[str] = []
    has_vehicles: bool = False


class NaveReorderPayload(BaseModel):
    ids: List[str]


class Event(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    guard: str
    type: str
    nave_id: Optional[str] = None
    nave_name: Optional[str] = None
    note: Optional[str] = None
    photo_path: Optional[str] = None
    turno_id: Optional[str] = None
    lat: Optional[float] = None
    lng: Optional[float] = None
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class EventCreate(BaseModel):
    guard: str
    type: str
    nave_id: Optional[str] = None
    nave_name: Optional[str] = None
    note: Optional[str] = None
    photo_path: Optional[str] = None
    turno_id: Optional[str] = None
    lat: Optional[float] = None
    lng: Optional[float] = None


class Incident(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    guard: str
    tipo: str
    nave_id: Optional[str] = None
    nave_name: Optional[str] = None
    description: str
    photo_path: Optional[str] = None
    turno_id: Optional[str] = None
    lat: Optional[float] = None
    lng: Optional[float] = None
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class IncidentCreate(BaseModel):
    guard: str
    tipo: str
    nave_id: Optional[str] = None
    nave_name: Optional[str] = None
    description: str
    photo_path: Optional[str] = None
    turno_id: Optional[str] = None
    lat: Optional[float] = None
    lng: Optional[float] = None


class Turno(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    guard: str
    service_name: str
    turno_tipo: Optional[str] = None  # dia | noche
    start_time: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    scheduled_end: Optional[datetime] = None
    end_time: Optional[datetime] = None
    status: str = "activo"  # activo | finalizado
    auto_finalizado: bool = False
    summary: Optional[dict] = None


class TurnoCreate(BaseModel):
    guard: str
    service_name: str
    turno_tipo: Optional[str] = None
    lat: Optional[float] = None
    lng: Optional[float] = None


class TurnoFinalizarPayload(BaseModel):
    lat: Optional[float] = None
    lng: Optional[float] = None


class Vehicle(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    nave_id: str
    tipo: str
    matricula: str
    zone: str = "linea"  # linea (cochera) | frente (aparcados)
    order: int = 0
    vandalizado: bool = False
    vandalizado_detalle: Optional[str] = None
    photo_path: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class VehicleCreate(BaseModel):
    nave_id: str
    tipo: str
    matricula: str
    zone: str = "linea"
    vandalizado: bool = False
    vandalizado_detalle: Optional[str] = None
    photo_path: Optional[str] = None
    guard: Optional[str] = None
    turno_id: Optional[str] = None


class VehicleUpdate(BaseModel):
    tipo: Optional[str] = None
    matricula: Optional[str] = None
    zone: Optional[str] = None
    vandalizado: Optional[bool] = None
    vandalizado_detalle: Optional[str] = None
    photo_path: Optional[str] = None
    guard: Optional[str] = None
    turno_id: Optional[str] = None


class ReorderPayload(BaseModel):
    nave_id: str
    zone: str
    ids: List[str]


class NaveCheck(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    nave_id: str
    turno_id: str
    item_name: str
    checked: bool = False
    checked_by: Optional[str] = None
    checked_at: Optional[datetime] = None


class Task(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    description: Optional[str] = None
    nave_id: Optional[str] = None
    nave_name: Optional[str] = None
    done: bool = False
    done_by: Optional[str] = None
    done_at: Optional[datetime] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class TaskCreate(BaseModel):
    title: str
    description: Optional[str] = None
    nave_id: Optional[str] = None
    nave_name: Optional[str] = None


# ---------------- Helpers ----------------
def clean(doc: dict) -> dict:
    if not doc:
        return doc
    doc.pop("_id", None)
    return doc


# ---------------- Routes ----------------
@api_router.get("/")
async def root():
    return {"message": "Control Guardia API"}


# --- Naves ---
@api_router.get("/naves", response_model=List[Nave])
async def list_naves():
    rows = await db.naves.find({}, {"_id": 0}).sort("order", 1).to_list(500)
    return [Nave(**r) for r in rows]


@api_router.post("/naves", response_model=Nave)
async def create_nave(payload: NaveCreate):
    count = await db.naves.count_documents({})
    obj = Nave(**payload.dict(), order=count)
    await db.naves.insert_one(obj.dict())
    return obj


@api_router.delete("/naves/{nave_id}")
async def delete_nave(nave_id: str):
    res = await db.naves.delete_one({"id": nave_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Nave no encontrada")
    return {"ok": True}


@api_router.post("/naves/reorder")
async def reorder_naves(payload: NaveReorderPayload):
    for idx, nid in enumerate(payload.ids):
        await db.naves.update_one({"id": nid}, {"$set": {"order": idx}})
    return {"ok": True}


# --- Servicios ---
@api_router.get("/services")
async def list_services():
    return [{"name": s} for s in SERVICES]


def compute_scheduled_end(start: datetime, turno_tipo: str) -> datetime:
    start_madrid = to_madrid(start)
    if turno_tipo == "dia":
        end = start_madrid.replace(hour=20, minute=0, second=0, microsecond=0)
        if end <= start_madrid:
            end += timedelta(days=1)
    else:  # noche
        tomorrow = start_madrid + timedelta(days=1)
        end = tomorrow.replace(hour=6, minute=0, second=0, microsecond=0)
    return end.astimezone(timezone.utc)


def is_weekend_madrid() -> bool:
    return now_madrid().weekday() in (5, 6)  # Sat=5, Sun=6


def validate_turno_tipo_for_today(turno_tipo: str) -> bool:
    if turno_tipo == "dia":
        return is_weekend_madrid()
    return turno_tipo == "noche"


@api_router.get("/turnos/opciones")
async def turno_opciones():
    """Opciones de turno disponibles hoy, según las reglas de Cándido Zamora:
    Día: solo sáb/dom 08:00-20:00. Noche: L-V 22:00-06:00, sáb/dom 20:30-06:00."""
    opciones = []
    if is_weekend_madrid():
        opciones.append({"tipo": "dia", "label": "Turno de Día", "horario": "08:00 - 20:00"})
        opciones.append({"tipo": "noche", "label": "Turno Nocturno", "horario": "20:30 - 06:00"})
    else:
        opciones.append({"tipo": "noche", "label": "Turno Nocturno", "horario": "22:00 - 06:00"})
    return opciones


async def _compute_turno_summary(turno_id: str, start_time, end_time: datetime) -> dict:
    events = await db.events.find({"turno_id": turno_id}, {"_id": 0}).to_list(3000)
    incidencias = await db.incidents.count_documents({"turno_id": turno_id})
    duracion_segundos = None
    if isinstance(start_time, datetime):
        st = start_time if start_time.tzinfo else start_time.replace(tzinfo=timezone.utc)
        duracion_segundos = (end_time - st).total_seconds()
    return {
        "total_eventos": len(events),
        "incidencias": incidencias,
        "llamadas_centralita": sum(1 for e in events if e.get("type") == "llamada_centralita"),
        "entradas_nave": sum(1 for e in events if e.get("type") == "entrada_nave"),
        "salidas_nave": sum(1 for e in events if e.get("type") == "salida_nave"),
        "descansos": sum(1 for e in events if e.get("type") == "descanso_inicio"),
        "chequeos": sum(1 for e in events if e.get("type") == "chequeo"),
        "duracion_segundos": duracion_segundos,
    }


async def _finalize_turno_doc(row: dict, auto: bool = False, lat: Optional[float] = None, lng: Optional[float] = None) -> dict:
    if row.get("status") == "finalizado":
        return row
    end_time = datetime.now(timezone.utc)
    summary = await _compute_turno_summary(row["id"], row.get("start_time"), end_time)
    update = {"end_time": end_time, "status": "finalizado", "summary": summary, "auto_finalizado": auto}
    await db.turnos.update_one({"id": row["id"]}, {"$set": update})
    await db.events.insert_one(Event(
        guard=row.get("guard", "—"), type="salida",
        note="Fin de turno (automático)" if auto else "Fin de turno",
        turno_id=row["id"], timestamp=end_time, lat=lat, lng=lng,
    ).dict())
    row.update(update)
    return row


async def _maybe_autofinalize(row: dict) -> dict:
    """Si la hora programada de fin de turno ya pasó (+ margen de cortesía) y el vigilante no
    finalizó manualmente, se finaliza automáticamente y queda guardado en el histórico de turnos."""
    if row.get("status") != "activo":
        return row
    now = datetime.now(timezone.utc)
    scheduled_end = row.get("scheduled_end")
    if isinstance(scheduled_end, datetime):
        se = scheduled_end if scheduled_end.tzinfo else scheduled_end.replace(tzinfo=timezone.utc)
        cutoff = se + timedelta(minutes=AUTOFINALIZE_GRACE_MINUTES)
        if now >= cutoff:
            return await _finalize_turno_doc(row, auto=True)
        return row
    # Turnos legacy sin scheduled_end (de antes de esta función): finalizar si llevan demasiado activos
    start_time = row.get("start_time")
    if isinstance(start_time, datetime):
        st = start_time if start_time.tzinfo else start_time.replace(tzinfo=timezone.utc)
        if now - st > timedelta(hours=LEGACY_TURNO_MAX_HOURS):
            return await _finalize_turno_doc(row, auto=True)
    return row


async def _migrate_legacy_turnos():
    """Limpieza única: turnos 'activo' creados antes de existir scheduled_end quedan huérfanos
    y bloquean que ese vigilante inicie un turno nuevo. Se finalizan automáticamente."""
    count = 0
    cursor = db.turnos.find({"status": "activo", "scheduled_end": {"$exists": False}}, {"_id": 0})
    async for row in cursor:
        await _finalize_turno_doc(row, auto=True)
        count += 1
    if count:
        logger.info(f"Migración: {count} turno(s) activo(s) legacy (sin scheduled_end) finalizados automáticamente")


async def _autofinalize_loop():
    while True:
        try:
            cursor = db.turnos.find({"status": "activo"}, {"_id": 0})
            async for row in cursor:
                await _maybe_autofinalize(row)
        except Exception as e:
            logger.warning(f"Error en autofinalize loop: {e}")
        await asyncio.sleep(120)


# --- Turnos (sesiones de trabajo) ---
@api_router.post("/turnos", response_model=Turno)
async def start_turno(payload: TurnoCreate):
    existing = await db.turnos.find_one({"guard": payload.guard, "status": "activo"}, {"_id": 0})
    if existing:
        existing = await _maybe_autofinalize(existing)
        if existing.get("status") == "activo":
            return Turno(**existing)

    turno_tipo = payload.turno_tipo or "noche"
    if turno_tipo not in TURNO_TIPOS:
        raise HTTPException(400, "Tipo de turno inválido")
    if not validate_turno_tipo_for_today(turno_tipo):
        raise HTTPException(400, "El turno de día solo está disponible sábado y domingo")

    start_time = datetime.now(timezone.utc)
    scheduled_end = compute_scheduled_end(start_time, turno_tipo)
    obj = Turno(
        guard=payload.guard, service_name=payload.service_name, turno_tipo=turno_tipo,
        start_time=start_time, scheduled_end=scheduled_end,
    )
    await db.turnos.insert_one(obj.dict())
    await db.events.insert_one(Event(
        guard=payload.guard, type="entrada", note="Inicio de turno",
        turno_id=obj.id, timestamp=start_time, lat=payload.lat, lng=payload.lng,
    ).dict())
    return obj


@api_router.get("/turnos/active")
async def get_active_turno(guard: str = Query(...)):
    row = await db.turnos.find_one({"guard": guard, "status": "activo"}, {"_id": 0})
    if not row:
        return None
    row = await _maybe_autofinalize(row)
    if row.get("status") != "activo":
        return None
    return Turno(**row)


@api_router.get("/turnos", response_model=List[Turno])
async def list_turnos(guard: Optional[str] = None, limit: int = 100):
    q = {}
    if guard:
        q["guard"] = guard
    rows = await db.turnos.find(q, {"_id": 0}).sort("start_time", -1).to_list(limit)
    return [Turno(**r) for r in rows]


@api_router.get("/turnos/{turno_id}", response_model=Turno)
async def get_turno(turno_id: str):
    row = await db.turnos.find_one({"id": turno_id}, {"_id": 0})
    if not row:
        raise HTTPException(404, "Turno no encontrado")
    return Turno(**row)


@api_router.post("/turnos/{turno_id}/finalizar", response_model=Turno)
async def finalizar_turno(turno_id: str, payload: Optional[TurnoFinalizarPayload] = Body(default=None)):
    row = await db.turnos.find_one({"id": turno_id}, {"_id": 0})
    if not row:
        raise HTTPException(404, "Turno no encontrado")
    lat = payload.lat if payload else None
    lng = payload.lng if payload else None
    row = await _finalize_turno_doc(row, auto=False, lat=lat, lng=lng)
    return Turno(**row)


# --- Vehículos ---
@api_router.get("/naves/{nave_id}/vehiculos", response_model=List[Vehicle])
async def list_vehicles(nave_id: str):
    rows = await db.vehicles.find({"nave_id": nave_id}, {"_id": 0}).sort("order", 1).to_list(500)
    return [Vehicle(**r) for r in rows]


async def _log_vandalizado_event(nave_id: str, guard: Optional[str], turno_id: Optional[str],
                                  tipo: str, matricula: str, vandalizado: bool,
                                  detalle: Optional[str], photo_path: Optional[str]):
    nave = await db.naves.find_one({"id": nave_id}, {"_id": 0})
    if vandalizado:
        note = f"{tipo} {matricula}" + (f": {detalle}" if detalle else "")
        ev_type = "vehiculo_vandalizado"
    else:
        note = f"{tipo} {matricula}: reparado / sin daños"
        ev_type = "vehiculo_reparado"
        photo_path = None
    ev = Event(
        guard=guard or "—", type=ev_type, nave_id=nave_id,
        nave_name=nave.get("name") if nave else None,
        note=note, photo_path=photo_path, turno_id=turno_id,
    )
    await db.events.insert_one(ev.dict())


@api_router.post("/vehiculos", response_model=Vehicle)
async def create_vehicle(payload: VehicleCreate):
    if payload.tipo not in VEHICLE_TIPOS:
        raise HTTPException(400, "Tipo de vehículo inválido")
    count = await db.vehicles.count_documents({"nave_id": payload.nave_id, "zone": payload.zone})
    data = payload.dict(exclude={"guard", "turno_id"})
    obj = Vehicle(**data, order=count)
    await db.vehicles.insert_one(obj.dict())
    if obj.vandalizado:
        await _log_vandalizado_event(
            obj.nave_id, payload.guard, payload.turno_id, obj.tipo, obj.matricula,
            True, obj.vandalizado_detalle, obj.photo_path,
        )
    return obj


@api_router.patch("/vehiculos/{vehicle_id}", response_model=Vehicle)
async def update_vehicle(vehicle_id: str, payload: VehicleUpdate):
    row = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    if not row:
        raise HTTPException(404, "Vehículo no encontrado")
    was_vandalizado = row.get("vandalizado", False)
    update = {k: v for k, v in payload.dict(exclude_unset=True, exclude={"guard", "turno_id"}).items()}
    update["updated_at"] = datetime.now(timezone.utc)
    await db.vehicles.update_one({"id": vehicle_id}, {"$set": update})
    row.update(update)
    if "vandalizado" in payload.dict(exclude_unset=True):
        new_vandalizado = row.get("vandalizado", False)
        if new_vandalizado or (was_vandalizado and not new_vandalizado):
            await _log_vandalizado_event(
                row["nave_id"], payload.guard, payload.turno_id, row.get("tipo"), row.get("matricula"),
                new_vandalizado, row.get("vandalizado_detalle"), row.get("photo_path"),
            )
    return Vehicle(**row)


@api_router.delete("/vehiculos/{vehicle_id}")
async def delete_vehicle(vehicle_id: str):
    res = await db.vehicles.delete_one({"id": vehicle_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Vehículo no encontrado")
    return {"ok": True}


@api_router.post("/vehiculos/reorder")
async def reorder_vehicles(payload: ReorderPayload):
    for idx, vid in enumerate(payload.ids):
        await db.vehicles.update_one(
            {"id": vid, "nave_id": payload.nave_id},
            {"$set": {"zone": payload.zone, "order": idx}},
        )
    return {"ok": True}


# --- Cajetines de verificación de nave (se reinician en cada turno nuevo) ---
@api_router.get("/naves/{nave_id}/checks")
async def get_nave_checks(nave_id: str, turno_id: str = Query(...)):
    nave = await db.naves.find_one({"id": nave_id}, {"_id": 0})
    items = (nave or {}).get("check_items", [])
    result = []
    for item in items:
        row = await db.nave_checks.find_one({"nave_id": nave_id, "turno_id": turno_id, "item_name": item}, {"_id": 0})
        if not row:
            row = {"nave_id": nave_id, "turno_id": turno_id, "item_name": item, "checked": False, "checked_by": None, "checked_at": None}
        result.append(row)
    return result


@api_router.post("/naves/{nave_id}/checks/{item_name}/toggle")
async def toggle_nave_check(nave_id: str, item_name: str, guard: str = Query(...), turno_id: str = Query(...)):
    nave = await db.naves.find_one({"id": nave_id}, {"_id": 0})
    valid_items = (nave or {}).get("check_items", [])
    if item_name not in valid_items:
        raise HTTPException(400, "Ítem no válido")
    row = await db.nave_checks.find_one({"nave_id": nave_id, "turno_id": turno_id, "item_name": item_name}, {"_id": 0})
    new_checked = not (row.get("checked") if row else False)
    now = datetime.now(timezone.utc)
    update = {
        "nave_id": nave_id,
        "turno_id": turno_id,
        "item_name": item_name,
        "checked": new_checked,
        "checked_by": guard if new_checked else None,
        "checked_at": now if new_checked else None,
    }
    await db.nave_checks.update_one(
        {"nave_id": nave_id, "turno_id": turno_id, "item_name": item_name},
        {"$set": update, "$setOnInsert": {"id": str(uuid.uuid4())}},
        upsert=True,
    )
    nave = await db.naves.find_one({"id": nave_id}, {"_id": 0})
    ev = Event(
        guard=guard, type="chequeo", nave_id=nave_id, nave_name=nave.get("name") if nave else None,
        turno_id=turno_id,
        note=f"{item_name}: {'Verificado' if new_checked else 'Desmarcado'}",
    )
    await db.events.insert_one(ev.dict())
    result = await db.nave_checks.find_one({"nave_id": nave_id, "turno_id": turno_id, "item_name": item_name}, {"_id": 0})
    result.pop("_id", None)
    return result


# --- Events / Control Horario ---
@api_router.get("/events", response_model=List[Event])
async def list_events(guard: Optional[str] = None, turno_id: Optional[str] = None, limit: int = 500):
    q = {}
    if guard:
        q["guard"] = guard
    if turno_id:
        q["turno_id"] = turno_id
    rows = await db.events.find(q, {"_id": 0}).sort("timestamp", -1).to_list(limit)
    return [Event(**r) for r in rows]


@api_router.post("/events", response_model=Event)
async def create_event(payload: EventCreate):
    if payload.type not in EVENT_TYPES:
        raise HTTPException(400, "Tipo de evento inválido")
    obj = Event(**payload.dict())
    await db.events.insert_one(obj.dict())
    return obj


@api_router.delete("/events/{event_id}")
async def delete_event(event_id: str):
    res = await db.events.delete_one({"id": event_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Evento no encontrado")
    return {"ok": True}


# --- Incidents ---
@api_router.get("/incidents", response_model=List[Incident])
async def list_incidents(guard: Optional[str] = None, turno_id: Optional[str] = None):
    q = {}
    if guard:
        q["guard"] = guard
    if turno_id:
        q["turno_id"] = turno_id
    rows = await db.incidents.find(q, {"_id": 0}).sort("timestamp", -1).to_list(500)
    return [Incident(**r) for r in rows]


@api_router.post("/incidents", response_model=Incident)
async def create_incident(payload: IncidentCreate):
    obj = Incident(**payload.dict())
    await db.incidents.insert_one(obj.dict())
    # also register a matching event
    ev = Event(
        guard=obj.guard,
        type="incidencia",
        nave_id=obj.nave_id,
        nave_name=obj.nave_name,
        turno_id=obj.turno_id,
        note=f"{obj.tipo}: {obj.description[:60]}",
        lat=obj.lat, lng=obj.lng,
    )
    await db.events.insert_one(ev.dict())
    return obj


# --- Tasks ---
@api_router.get("/tasks", response_model=List[Task])
async def list_tasks():
    rows = await db.tasks.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return [Task(**r) for r in rows]


@api_router.post("/tasks", response_model=Task)
async def create_task(payload: TaskCreate):
    obj = Task(**payload.dict())
    await db.tasks.insert_one(obj.dict())
    return obj


@api_router.post("/tasks/{task_id}/toggle", response_model=Task)
async def toggle_task(task_id: str, guard: str = Query(...)):
    row = await db.tasks.find_one({"id": task_id}, {"_id": 0})
    if not row:
        raise HTTPException(404, "Tarea no encontrada")
    new_done = not row.get("done", False)
    update = {
        "done": new_done,
        "done_by": guard if new_done else None,
        "done_at": datetime.now(timezone.utc) if new_done else None,
    }
    await db.tasks.update_one({"id": task_id}, {"$set": update})
    row.update(update)
    if new_done:
        ev = Event(
            guard=guard, type="tarea", nave_id=row.get("nave_id"), nave_name=row.get("nave_name"),
            note=f"Completada: {row.get('title')}"
        )
        await db.events.insert_one(ev.dict())
    return Task(**row)


@api_router.delete("/tasks/{task_id}")
async def delete_task(task_id: str):
    res = await db.tasks.delete_one({"id": task_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Tarea no encontrada")
    return {"ok": True}


# --- File upload (photos) ---
@api_router.post("/upload")
async def upload_file(file: UploadFile = File(...), guard: str = Form("anon")):
    ext = (file.filename or "photo.jpg").split(".")[-1].lower()
    if ext not in ("jpg", "jpeg", "png", "webp", "heic"):
        ext = "jpg"
    safe_guard = "".join(c for c in guard if c.isalnum() or c in "-_") or "anon"
    obj_path = f"{APP_NAME}/uploads/{safe_guard}/{uuid.uuid4().hex}.{ext}"
    data = await file.read()
    content_type = file.content_type or f"image/{ext}"
    await run_in_threadpool(put_object, obj_path, data, content_type)
    return {"path": obj_path, "size": len(data)}


@api_router.get("/files/{full_path:path}")
async def get_file(full_path: str):
    try:
        data, ct = await run_in_threadpool(get_object, full_path)
    except Exception:
        raise HTTPException(404, "Archivo no encontrado")
    return Response(content=data, media_type=ct)


# --- Export ---
def _format_timestamp(ts) -> str:
    if isinstance(ts, str):
        try:
            ts = datetime.fromisoformat(ts)
        except Exception:
            return ts
    if isinstance(ts, datetime):
        return to_madrid(ts).strftime("%d/%m/%Y %H:%M:%S")
    return str(ts)


EVENT_LABELS = {
    "entrada": "Fichar Entrada",
    "salida": "Fichar Salida",
    "ronda_inicio": "Inicio Ronda",
    "ronda_fin": "Fin Ronda",
    "tarea": "Tarea",
    "incidencia": "Incidencia",
    "descanso_inicio": "Inicio Descanso",
    "descanso_fin": "Fin Descanso",
    "entrada_nave": "Entrada a Nave",
    "salida_nave": "Salida de Nave",
    "llamada_centralita": "Llamada Centralita",
    "chequeo": "Chequeo",
    "accion_nave": "Acción de Nave",
    "vehiculo_vandalizado": "Vehículo Vandalizado",
    "vehiculo_reparado": "Vehículo Reparado / Sin Daños",
}


def _format_duration(seconds: Optional[float]) -> str:
    if not seconds or seconds < 0:
        return "—"
    h = int(seconds // 3600)
    m = int((seconds % 3600) // 60)
    return f"{h}h {m:02d}min"


# --- Identidad de marca ASERGRUP para exportaciones PDF/Excel ---
ASSETS_DIR = Path(__file__).resolve().parent.parent / "assets_src"
LOGO_DARK_PATH = ASSETS_DIR / "logo_horizontal_dark.png"   # logo + wordmark sobre fondo oscuro (cabecera)
LOGO_LIGHT_PATH = ASSETS_DIR / "logo_full_light.png"        # logo + wordmark sobre fondo claro

BRAND_DARK = colors.HexColor("#0D1117")
BRAND_RED = colors.HexColor("#E53935")
BRAND_GRAY = colors.HexColor("#5A606C")
BRAND_LIGHT = colors.HexColor("#F2F4F7")
BRAND_GRID = colors.HexColor("#D8DCE1")

PDF_MARGIN = 40
PDF_CONTENT_WIDTH = A4[0] - (PDF_MARGIN * 2)  # ~515pt disponibles


def _pdf_footer(canvas, doc):
    canvas.saveState()
    canvas.setStrokeColor(BRAND_GRID)
    canvas.setLineWidth(0.6)
    canvas.line(PDF_MARGIN, 34, A4[0] - PDF_MARGIN, 34)
    canvas.setFont("Helvetica", 7.5)
    canvas.setFillColor(BRAND_GRAY)
    canvas.drawString(PDF_MARGIN, 22, "ASERGRUP · Seguridad Integral — documento de uso interno")
    canvas.drawRightString(A4[0] - PDF_MARGIN, 22, f"Página {doc.page}")
    canvas.restoreState()


def _pdf_header_block():
    """Banda de cabecera oscura con el logo + una fina línea roja de acento."""
    header_tbl = Table([[RLImage(str(LOGO_DARK_PATH), width=190, height=58)]], colWidths=[PDF_CONTENT_WIDTH])
    header_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), BRAND_DARK),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 14),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 14),
    ]))
    accent = Table([[""]], colWidths=[PDF_CONTENT_WIDTH], rowHeights=[3])
    accent.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), BRAND_RED)]))
    return [header_tbl, accent]


def _pdf_section_header(text: str):
    """Título de sección en mayúsculas con una línea roja debajo, estilo informe corporativo."""
    p = Paragraph(f"<b>{text}</b>", ParagraphStyle(
        "SecHead", fontName="Helvetica-Bold", fontSize=11, textColor=BRAND_DARK, leading=13,
    ))
    tbl = Table([[p]], colWidths=[PDF_CONTENT_WIDTH])
    tbl.setStyle(TableStyle([
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("LINEBELOW", (0, 0), (-1, -1), 1.3, BRAND_RED),
    ]))
    return tbl


def _pdf_kpi_row(stats: list):
    """Fila de tarjetas resumen (número grande + etiqueta), tipo dashboard ejecutivo."""
    n = len(stats)
    col_w = PDF_CONTENT_WIDTH / n
    cells = []
    for number, label in stats:
        cells.append(Paragraph(
            f'<font color="#E53935" size="20"><b>{number}</b></font><br/>'
            f'<font color="#5A606C" size="8">{label}</font>',
            ParagraphStyle("Kpi", alignment=1, leading=22),
        ))
    tbl = Table([cells], colWidths=[col_w] * n)
    style = [
        ("BACKGROUND", (0, 0), (-1, -1), BRAND_LIGHT),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("BOX", (0, 0), (-1, -1), 0.6, BRAND_GRID),
    ]
    for i in range(n - 1):
        style.append(("LINEAFTER", (i, 0), (i, 0), 0.6, BRAND_GRID))
    tbl.setStyle(TableStyle(style))
    return tbl


def _xl_style_sheet(ws, headers: list, header_row: int, thin_border: Border):
    """Aplica cabecera oscura de marca, bordes y ancho de fila a una hoja."""
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.fill = PatternFill(start_color="0D1117", end_color="0D1117", fill_type="solid")
        c.font = Font(color="F2F4F7", bold=True, size=11)
        c.border = thin_border
    ws.row_dimensions[header_row].height = 20


def _xl_add_header_band(ws, n_cols: int):
    """Cabecera oscura con el logo ASERGRUP + una fina línea roja de acento, tipo membrete."""
    last_col = chr(64 + n_cols)
    ws.merge_cells(f"A1:{last_col}1")
    ws.row_dimensions[1].height = 46
    dark_fill = PatternFill(start_color="0D1117", end_color="0D1117", fill_type="solid")
    for col in range(1, n_cols + 1):
        ws.cell(row=1, column=col).fill = dark_fill
    if LOGO_DARK_PATH.exists():
        img = XLImage(str(LOGO_DARK_PATH))
        img.width, img.height = 176, 54
        ws.add_image(img, "A1")

    ws.merge_cells(f"A2:{last_col}2")
    ws.row_dimensions[2].height = 4
    red_fill = PatternFill(start_color="E53935", end_color="E53935", fill_type="solid")
    for col in range(1, n_cols + 1):
        ws.cell(row=2, column=col).fill = red_fill

    ws.row_dimensions[3].height = 8  # respiro antes del contenido


@api_router.get("/export/excel")
async def export_excel(guard: Optional[str] = None, turno_id: Optional[str] = None):
    q = {}
    if guard:
        q["guard"] = guard
    if turno_id:
        q["turno_id"] = turno_id
    rows = await db.events.find(q, {"_id": 0}).sort("timestamp", 1).to_list(5000)

    turno = await db.turnos.find_one({"id": turno_id}, {"_id": 0}) if turno_id else None

    wb = Workbook()
    ws = wb.active
    ws.title = "Control Horario"
    ws.sheet_properties.tabColor = "E53935"

    thin = Side(style="thin", color="D8DCE1")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    highlight_fill = PatternFill(start_color="FDE0DF", end_color="FDE0DF", fill_type="solid")
    highlight_font = Font(color="B0221E", bold=True)
    zebra_fill = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")

    headers = ["Fecha/Hora", "Vigilante", "Evento", "Nave", "Notas", "Ubicación"]
    _xl_add_header_band(ws, len(headers))
    row_i = 4

    if turno:
        info = [
            ("Vigilante", turno.get("guard", "")),
            ("Servicio", turno.get("service_name", "")),
            ("Turno", "Día" if turno.get("turno_tipo") == "dia" else "Nocturno"),
            ("Día de servicio", _format_timestamp(turno.get("start_time")).split(" ")[0]),
            ("Inicio", _format_timestamp(turno.get("start_time"))),
            ("Fin", _format_timestamp(turno.get("end_time")) if turno.get("end_time") else "En curso"),
            ("Horas trabajadas", _format_duration((turno.get("summary") or {}).get("duracion_segundos"))),
        ]
        for label, value in info:
            lc = ws.cell(row=row_i, column=1, value=label)
            lc.font = Font(bold=True, color="0D1117")
            lc.fill = zebra_fill
            lc.border = thin_border
            vc = ws.cell(row=row_i, column=2, value=value)
            vc.fill = zebra_fill
            vc.border = thin_border
            row_i += 1
        row_i += 1

    header_row = row_i
    _xl_style_sheet(ws, headers, header_row, thin_border)
    row_i += 1

    for idx, r in enumerate(rows):
        is_highlight = r.get("type") in ("incidencia", "llamada_centralita")
        lat, lng = r.get("lat"), r.get("lng")
        values = [
            _format_timestamp(r.get("timestamp")),
            r.get("guard", ""),
            EVENT_LABELS.get(r.get("type", ""), r.get("type", "")),
            r.get("nave_name") or "",
            r.get("note") or "",
            "Ver mapa" if (lat is not None and lng is not None) else "",
        ]
        for col, v in enumerate(values, start=1):
            c = ws.cell(row=row_i, column=col, value=v)
            c.border = thin_border
            if col == 6 and v:
                c.hyperlink = f"https://maps.google.com/?q={lat},{lng}"
                c.font = Font(color="1155CC", underline="single")
            if is_highlight:
                c.fill = highlight_fill
                if col != 6:
                    c.font = highlight_font
            elif idx % 2 == 1:
                c.fill = zebra_fill
        row_i += 1

    last_data_row = row_i - 1
    if last_data_row >= header_row:
        last_col_letter = chr(64 + len(headers))
        ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{last_data_row}"
    ws.freeze_panes = f"A{header_row + 1}"

    total_cell = ws.cell(row=row_i + 1, column=1, value=f"Total de registros: {len(rows)}")
    total_cell.font = Font(bold=True, italic=True, color="5A606C", size=9)

    for col_idx, width in enumerate([22, 20, 22, 24, 40, 14], start=1):
        ws.column_dimensions[chr(64 + col_idx)].width = width

    if turno and turno.get("service_name") == "Cándido Zamora":
        ws2 = wb.create_sheet("Vehículos Cerámicas")
        ws2.sheet_properties.tabColor = "E53935"
        v_headers = ["Zona", "Tipo", "Matrícula", "Estado", "Detalle"]
        _xl_add_header_band(ws2, len(v_headers))
        v_header_row = 4
        _xl_style_sheet(ws2, v_headers, v_header_row, thin_border)
        v_row = v_header_row + 1

        ceramicas = await db.naves.find_one({"name": "Cerámicas"}, {"_id": 0})
        if ceramicas:
            vehicles = await db.vehicles.find({"nave_id": ceramicas["id"]}, {"_id": 0}).sort("order", 1).to_list(200)
            for idx, v in enumerate(vehicles):
                values = [
                    "Línea" if v.get("zone") == "linea" else "Frente",
                    v.get("tipo"), v.get("matricula"),
                    "VANDALIZADO" if v.get("vandalizado") else "OK",
                    v.get("vandalizado_detalle") or "",
                ]
                for col, val in enumerate(values, start=1):
                    c = ws2.cell(row=v_row, column=col, value=val)
                    c.border = thin_border
                    if v.get("vandalizado"):
                        c.fill = highlight_fill
                        c.font = highlight_font
                    elif idx % 2 == 1:
                        c.fill = zebra_fill
                v_row += 1

        if v_row - 1 >= v_header_row:
            ws2.auto_filter.ref = f"A{v_header_row}:E{v_row - 1}"
        ws2.freeze_panes = f"A{v_header_row + 1}"
        for col_idx, width in enumerate([12, 14, 16, 16, 30], start=1):
            ws2.column_dimensions[chr(64 + col_idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"control_horario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )



@api_router.get("/export/pdf")
async def export_pdf(guard: Optional[str] = None, turno_id: Optional[str] = None):
    q = {}
    if guard:
        q["guard"] = guard
    if turno_id:
        q["turno_id"] = turno_id
    rows = await db.events.find(q, {"_id": 0}).sort("timestamp", 1).to_list(5000)
    incidents = await db.incidents.find(q, {"_id": 0}).sort("timestamp", 1).to_list(500) if turno_id else []

    turno = await db.turnos.find_one({"id": turno_id}, {"_id": 0}) if turno_id else None

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4, title="Control Diario Asergrup",
        leftMargin=PDF_MARGIN, rightMargin=PDF_MARGIN, topMargin=36, bottomMargin=46,
    )
    styles = getSampleStyleSheet()
    story = []
    story.extend(_pdf_header_block())
    story.append(Spacer(1, 14))

    story.append(Paragraph("INFORME DE CONTROL DE SERVICIO", ParagraphStyle(
        "Kicker", fontName="Helvetica-Bold", fontSize=10, textColor=BRAND_RED,
    )))
    story.append(Paragraph(
        f"Generado: {to_madrid(datetime.now(timezone.utc)).strftime('%d/%m/%Y %H:%M')}",
        ParagraphStyle("Gen", fontSize=8.5, textColor=BRAND_GRAY),
    ))
    story.append(Spacer(1, 10))

    if turno:
        turno_label = "Turno de Día" if turno.get("turno_tipo") == "dia" else "Turno Nocturno"
        dia_servicio = _format_timestamp(turno.get("start_time")).split(" ")[0]
        horas = _format_duration((turno.get("summary") or {}).get("duracion_segundos"))
        info_data = [
            ["Vigilante", turno.get("guard", ""), "Servicio", turno.get("service_name", "")],
            ["Turno", turno_label, "Día de servicio", dia_servicio],
            ["Inicio", _format_timestamp(turno.get("start_time")), "Fin",
             _format_timestamp(turno.get("end_time")) if turno.get("end_time") else "En curso"],
            ["Horas trabajadas", horas, "", ""],
        ]
        info_tbl = Table(info_data, colWidths=[PDF_CONTENT_WIDTH * 0.16, PDF_CONTENT_WIDTH * 0.34,
                                                 PDF_CONTENT_WIDTH * 0.16, PDF_CONTENT_WIDTH * 0.34])
        info_tbl.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
            ("TEXTCOLOR", (0, 0), (0, -1), BRAND_GRAY),
            ("TEXTCOLOR", (2, 0), (2, -1), BRAND_GRAY),
            ("BACKGROUND", (0, 0), (-1, -1), BRAND_LIGHT),
            ("BOX", (0, 0), (-1, -1), 0.6, BRAND_GRID),
            ("INNERGRID", (0, 0), (-1, -1), 0.4, BRAND_GRID),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("SPAN", (1, 3), (3, 3)),
        ]))
        story.append(info_tbl)
    else:
        subtitle = f"Vigilante: {guard}" if guard else "Todos los vigilantes"
        story.append(Paragraph(subtitle, styles["Normal"]))
    story.append(Spacer(1, 14))

    n_incidencias = sum(1 for r in rows if r.get("type") == "incidencia")
    n_llamadas = sum(1 for r in rows if r.get("type") == "llamada_centralita")
    story.append(_pdf_kpi_row([
        (len(rows), "EVENTOS TOTALES"),
        (n_incidencias, "INCIDENCIAS"),
        (n_llamadas, "LLAMADAS CENTRALITA"),
    ]))
    story.append(Spacer(1, 16))

    story.append(_pdf_section_header("HISTÓRICO DE NOVEDADES"))
    story.append(Spacer(1, 5))
    story.append(Paragraph(
        "Las incidencias y llamadas de centralita aparecen destacadas en rojo.",
        ParagraphStyle("Note", fontSize=8, textColor=BRAND_GRAY),
    ))
    story.append(Spacer(1, 6))

    data = [["Fecha/Hora", "Vigilante", "Evento", "Nave", "Notas"]]
    highlight_rows = []
    note_style = ParagraphStyle("NoteCell", fontSize=8, leading=10)
    for r in rows:
        if r.get("type") in ("incidencia", "llamada_centralita"):
            highlight_rows.append(len(data))
        note_text = (r.get("note") or "")[:60]
        lat, lng = r.get("lat"), r.get("lng")
        if lat is not None and lng is not None:
            note_cell = Paragraph(
                f'{note_text} <link href="https://maps.google.com/?q={lat},{lng}" color="#E53935"><u>📍 Ver mapa</u></link>',
                note_style,
            )
        else:
            note_cell = note_text
        data.append([
            _format_timestamp(r.get("timestamp")),
            r.get("guard", ""),
            EVENT_LABELS.get(r.get("type", ""), r.get("type", "")),
            r.get("nave_name") or "-",
            note_cell,
        ])
    if len(data) == 1:
        data.append(["Sin registros", "", "", "", ""])

    tbl_style = [
        ("BACKGROUND", (0, 0), (-1, 0), BRAND_DARK),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#F2F4F7")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.4, BRAND_GRID),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [BRAND_LIGHT, colors.white]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]
    for r_idx in highlight_rows:
        tbl_style.append(("BACKGROUND", (0, r_idx), (-1, r_idx), colors.HexColor("#FDE0DF")))
        tbl_style.append(("TEXTCOLOR", (0, r_idx), (-1, r_idx), colors.HexColor("#B0221E")))
        tbl_style.append(("FONTNAME", (0, r_idx), (-1, r_idx), "Helvetica-Bold"))

    tbl = Table(data, repeatRows=1, colWidths=[95, 70, 80, 90, 180])
    tbl.setStyle(TableStyle(tbl_style))
    story.append(tbl)

    if turno and incidents:
        story.append(Spacer(1, 16))
        story.append(_pdf_section_header("DETALLE DE INCIDENCIAS"))
        story.append(Spacer(1, 6))
        for inc in incidents:
            lat, lng = inc.get("lat"), inc.get("lng")
            map_link = (
                f' <font color="#E53935"><link href="https://maps.google.com/?q={lat},{lng}"><u>📍 Ver mapa</u></link></font>'
                if lat is not None and lng is not None else ""
            )
            story.append(Paragraph(
                f'<font color="#E53935">●</font> <b>{_format_timestamp(inc.get("timestamp"))}</b> — '
                f'{inc.get("tipo")} ({inc.get("nave_name") or "sin nave"}): {inc.get("description")}{map_link}',
                ParagraphStyle("IncItem", fontSize=9, leading=13, spaceAfter=4),
            ))
        story.append(Spacer(1, 6))

    if turno and turno.get("service_name") == "Cándido Zamora":
        ceramicas = await db.naves.find_one({"name": "Cerámicas"}, {"_id": 0})
        if ceramicas:
            vehicles = await db.vehicles.find({"nave_id": ceramicas["id"]}, {"_id": 0}).sort("order", 1).to_list(200)
            if vehicles:
                story.append(Spacer(1, 16))
                story.append(_pdf_section_header("RESUMEN DE VEHÍCULOS — NAVE CERÁMICAS"))
                story.append(Spacer(1, 6))
                vdata = [["Zona", "Tipo", "Matrícula", "Estado", "Detalle"]]
                vand_rows = []
                for v in vehicles:
                    if v.get("vandalizado"):
                        vand_rows.append(len(vdata))
                    vdata.append([
                        "Línea" if v.get("zone") == "linea" else "Frente",
                        v.get("tipo"), v.get("matricula"),
                        "VANDALIZADO" if v.get("vandalizado") else "OK",
                        (v.get("vandalizado_detalle") or "")[:40],
                    ])
                vstyle = [
                    ("BACKGROUND", (0, 0), (-1, 0), BRAND_DARK),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#F2F4F7")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.4, BRAND_GRID),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [BRAND_LIGHT, colors.white]),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]
                for r_idx in vand_rows:
                    vstyle.append(("BACKGROUND", (0, r_idx), (-1, r_idx), colors.HexColor("#FDE0DF")))
                    vstyle.append(("TEXTCOLOR", (0, r_idx), (-1, r_idx), colors.HexColor("#B0221E")))
                    vstyle.append(("FONTNAME", (0, r_idx), (-1, r_idx), "Helvetica-Bold"))
                vtbl = Table(vdata, repeatRows=1, colWidths=[65, 80, 90, 90, 190])
                vtbl.setStyle(TableStyle(vstyle))
                story.append(vtbl)

    doc.build(story, onFirstPage=_pdf_footer, onLaterPages=_pdf_footer)
    buf.seek(0)
    filename = f"control_horario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )



# ---------------- App wiring ----------------
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


async def seed_defaults():
    # --- Cerámicas: nave principal con vehículos ---
    ceramicas = await db.naves.find_one({"name": "Cerámicas"})
    ceramicas_config = {
        "has_access_buttons": True,
        "check_items": ["Luces traseras", "Luces camiones"],
        "custom_actions": [],
        "has_vehicles": True,
        "service_name": "Cándido Zamora",
    }
    if not ceramicas:
        nave = Nave(name="Cerámicas", notes="Nave principal del servicio", **ceramicas_config)
        await db.naves.insert_one(nave.dict())
        ceramicas_id = nave.id

        linea_data = [
            ("Camión", "8536 GVP", False, None),
            ("Camión", "5064 KHZ", False, None),
            ("Camión", "3847 MMX", False, None),
            ("Camión", "7126 FZB", False, None),
            ("Camión", "7650 FMC", False, None),
            ("Camión", "4613 HMS", False, None),
            ("Camión", "4849 HTV", False, None),
            ("Camión", "8458 JHB", False, None),
            ("Grúa", "6207 DGG", False, None),
            ("Grúa", "2507 DZM", False, None),
            ("Grúa", "3192 FXJ", False, None),
            ("Grúa", "2103 BSS", False, None),
            ("Grúa", "8327 DTW", False, None),
            ("Grúa", "9004 DJT", True, "Ambas cerraduras"),
            ("Contenedor", "9425 FLS", False, None),
            ("Contenedor", "6754 KTM", False, None),
            ("Contenedor", "4742 MHH", False, None),
        ]
        frente_data = [
            ("Grúa", "9859 CNS", False, None),
            ("Grúa", "TO 5990 AD", True, "Cerradura Copiloto"),
            ("Grúa", "5573 BHX", True, "Cerradura Copiloto"),
            ("Grúa", "3882 CCP", True, "Cerradura Conductor"),
        ]
        for idx, (tipo, matricula, vand, detalle) in enumerate(linea_data):
            v = Vehicle(nave_id=ceramicas_id, tipo=tipo, matricula=matricula, zone="linea", order=idx,
                        vandalizado=vand, vandalizado_detalle=detalle)
            await db.vehicles.insert_one(v.dict())
        for idx, (tipo, matricula, vand, detalle) in enumerate(frente_data):
            v = Vehicle(nave_id=ceramicas_id, tipo=tipo, matricula=matricula, zone="frente", order=idx,
                        vandalizado=vand, vandalizado_detalle=detalle)
            await db.vehicles.insert_one(v.dict())
        logger.info("Nave Cerámicas + vehículos sembrados correctamente")
    else:
        # Migración: asegurar campos de configuración en nave ya existente
        await db.naves.update_one({"id": ceramicas["id"]}, {"$set": {**ceramicas_config, "order": 0}})
        # Migración: corregir tipo/matrícula del vehículo con matrícula antigua
        await db.vehicles.update_many(
            {"nave_id": ceramicas["id"], "matricula": "5990 AD"},
            {"$set": {"tipo": "Grúa", "matricula": "TO 5990 AD"}},
        )
        await db.vehicles.update_many(
            {"nave_id": ceramicas["id"], "tipo": "Grúa TO"},
            {"$set": {"tipo": "Grúa"}},
        )

    # --- Resto de naves / puntos de control de Cándido Zamora ---
    otras_naves = [
        {
            "name": "Grúas", "notes": "Segunda nave principal",
            "has_access_buttons": True, "check_items": ["Luces torre", "Luces grúa"],
            "custom_actions": [], "has_vehicles": False,
        },
        {
            "name": "Oficinas", "notes": None,
            "has_access_buttons": False, "check_items": [],
            "custom_actions": ["Revisión de luces", "Cerrada", "Abierta"], "has_vehicles": False,
        },
        {
            "name": "PP3", "notes": "Rejas verdes frente a grúas",
            "has_access_buttons": True, "check_items": [],
            "custom_actions": ["Revisado"], "has_vehicles": False,
        },
        {
            "name": "Eólica", "notes": "Frente a Hierros Villaverde",
            "has_access_buttons": True, "check_items": [],
            "custom_actions": ["Revisión luces"], "has_vehicles": False,
        },
        {
            "name": "Nave Camino Agrícola", "notes": None,
            "has_access_buttons": True, "check_items": [],
            "custom_actions": [], "has_vehicles": False,
        },
        {
            "name": "Camino Agrícola", "notes": None,
            "has_access_buttons": False, "check_items": [],
            "custom_actions": ["Paso por revisión"], "has_vehicles": False,
        },
        {
            "name": "Camino Eólica-Grúas", "notes": "Camino de tierra de la gasolinera",
            "has_access_buttons": False, "check_items": [],
            "custom_actions": ["Paso por revisión"], "has_vehicles": False,
        },
    ]
    for order_idx, cfg in enumerate(otras_naves, start=1):
        exists = await db.naves.find_one({"name": cfg["name"]})
        if exists:
            await db.naves.update_one(
                {"id": exists["id"]},
                {"$set": {**cfg, "service_name": "Cándido Zamora", "order": exists.get("order", order_idx)}},
            )
            continue
        nave = Nave(service_name="Cándido Zamora", order=order_idx, **cfg)
        await db.naves.insert_one(nave.dict())
    logger.info("Naves adicionales de Cándido Zamora sembradas/actualizadas correctamente")


@app.on_event("startup")
async def on_startup():
    await db.connect(DATABASE_URL)
    logger.info("Postgres (Supabase) pool initialized")
    try:
        await run_in_threadpool(init_storage)
        logger.info("Object storage initialized")
    except Exception as e:
        logger.warning(f"Object storage init failed: {e}")
    try:
        await seed_defaults()
    except Exception as e:
        logger.warning(f"Seed defaults failed: {e}")
    try:
        await _migrate_legacy_turnos()
    except Exception as e:
        logger.warning(f"Legacy turno migration failed: {e}")
    asyncio.create_task(_autofinalize_loop())


@app.on_event("shutdown")
async def shutdown_db_client():
    await db.close()
